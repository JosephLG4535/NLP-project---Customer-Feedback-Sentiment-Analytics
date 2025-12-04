import pandas as pd 
import openpyxl 
import os 
# open new workbook & remove default sheet 
master_df = openpyxl.Workbook() 
master_df.remove(master_df.active) 
dir = os.path.join(os.getcwd(), "cleaned_xlsx") 
files = ['blinkit_customer_feedback.xlsx','blinkit_customers.xlsx','blinkit_delivery_performance.xlsx', 
         'blinkit_inventory.xlsx','blinkit_marketing_performance.xlsx','blinkit_order_items.xlsx', 
         'blinkit_orders.xlsx','blinkit_products.xlsx'] 
for file in files: 
     file_path = os.path.join(dir, file) 
     wb = openpyxl.load_workbook(file_path)
     for sheet in wb.sheetnames: #only one sheet in each file 
                                 #therefore master data dont have to specify which file the sheet is from 
              new_sheetname = sheet.replace('blinkit_','') 
              ws = wb[sheet] 
              new_ws = master_df.create_sheet(f'{new_sheetname}') 
              for row in ws.iter_rows(values_only=True): 
                   new_ws.append(row) 
save_path = os.path.join(dir, "consolidated.xlsx") 
master_df.save(save_path)